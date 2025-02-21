import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import tempfile
import webbrowser
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from matplotlib import use
use('TkAgg')

# ------------------------------------------------------------
# GLOBAL SETTINGS / DATAFRAMES
# ------------------------------------------------------------
DATABASE_FILE = "database.xlsx"

# Columns for each sheet
PROJECT_COLUMNS = [
    'ProjectID', 'ProjectName', 'Notes', 
    'ElectricalProgress', 'SSProgress', 'PlumbingProgress', 'ACProgress',
    'Wall Tiles Progress', 'Wall Partion Progress', 'floortiles progress',
    'Ceiling progress', 'Furniture Progress', 'FAFF Progress', 'Fire Suppersion',
    'IT', 'Signage', 'External Work', 'FireSuppressionProgress', 'OverallProgress',
    'Constraction Progress', 'ColdRoom Progress', 'Equipment Progress',
]

TASK_COLUMNS = [
    'TaskID', 'ProjectID', 'TaskName', 'Duration', 'Weight', 'Progress', 
    'ParentTaskID', 'Category', 'PendingItems'  # Added PendingItems
]

ORDER_COLUMNS = [
    'OrderID', 'ProjectID', 'Company', 'ItemCategory', 'OrderStatus', 'LPOStatus', 
    'InvoiceCopyPath', 'InvoiceStatus', 'MissingItems', 'DeliveryDate', 'InstallationDate'
]

# Sub-progress categories for tasks -> project subprogress
TASK_SUBCATEGORIES = {
    'Electrical': 'ElectricalProgress',
    'S/S': 'SSProgress',
    'Plumbing': 'PlumbingProgress',
    'AC': 'ACProgress',
    'Wall Tiles': 'Wall Tiles Progress',
    'Wall Partition': 'Wall Partion Progress',
    'Floor Tiles': 'floortiles progress',
    'Ceiling': 'Ceiling progress',
    'Furniture': 'Furniture Progress',
    'FAFF': 'FAFF Progress',
    'Fire Suppression': 'Fire Suppersion',
    'IT': 'IT',
    'Signage': 'Signage',
    'External Work': 'External Work',
    'Fire Suppression2': 'FireSuppressionProgress',
    'Constraction': 'Constraction Progress',
    'Cold Room': 'ColdRoom Progress',
    'Equipment': 'Equipment Progress',
}

# DataFrames in memory
PENDING_WORK_COLUMNS = ['PendingID', 'TaskID', 'ProjectID', 'Description', 'Status', 'DueDate']

pending_work_df = pd.DataFrame(columns=PENDING_WORK_COLUMNS)

projects_df = pd.DataFrame(columns=PROJECT_COLUMNS)
tasks_df = pd.DataFrame(columns=TASK_COLUMNS)
orders_df = pd.DataFrame(columns=ORDER_COLUMNS)

# Predefined reference data for orders
COMPANY_NAMES = [
    "No Company Selected", "Al motqeen", "Oriantal", "Himalya", "Richline",
    "Kain", "Al jaz", "A3", "Blue Rhain", "Wize guys", "Eco Air", "Tripode", "Metre"
]
ITEM_CATEGORIES = ["S/S", "Furniture", "Equipment", "Signage", "Fire Suppression", "Cold Room"]
ORDER_STATUSES = ["Ordered", "Not Ordered"]
LPO_STATUSES = ["LPO Received", "Pending", "LPO Pending"]
INVOICE_STATUSES = ["Not Submitted", "25%", "50%", "100%"]

# ------------------------------------------------------------
# LOADING / SAVING DATA
# ------------------------------------------------------------
def load_data():
    global projects_df, tasks_df, orders_df, pending_work_df
    try:
        projects_df = pd.read_excel(DATABASE_FILE, sheet_name='Projects')
    except FileNotFoundError:
        projects_df = pd.DataFrame(columns=PROJECT_COLUMNS)
    except Exception as e:
        print(f"Error loading Projects sheet: {e}")
        projects_df = pd.DataFrame(columns=PROJECT_COLUMNS)

    try:
        tasks_df = pd.read_excel(DATABASE_FILE, sheet_name='Tasks')
    except FileNotFoundError:
        tasks_df = pd.DataFrame(columns=TASK_COLUMNS)
    except Exception as e:
        print(f"Error loading Tasks sheet: {e}")
        tasks_df = pd.DataFrame(columns=TASK_COLUMNS)

    try:
        orders_df = pd.read_excel(DATABASE_FILE, sheet_name='Orders')
    except FileNotFoundError:
        orders_df = pd.DataFrame(columns=ORDER_COLUMNS)
    except Exception as e:
        print(f"Error loading Orders sheet: {e}")
        orders_df = pd.DataFrame(columns=ORDER_COLUMNS)

    try:
        pending_work_df = pd.read_excel(DATABASE_FILE, sheet_name='PendingWork')
    except FileNotFoundError:
        pending_work_df = pd.DataFrame(columns=PENDING_WORK_COLUMNS)
    except Exception as e:
        print(f"Error loading Pending Work sheet: {e}")
        pending_work_df = pd.DataFrame(columns=PENDING_WORK_COLUMNS)

    # Ensure all columns exist for each dataframe
    for col in PROJECT_COLUMNS:
        if col not in projects_df.columns:
            if 'Progress' in col:
                projects_df[col] = 0
            else:
                projects_df[col] = ""

    for col in TASK_COLUMNS:
        if col not in tasks_df.columns:
            if col in ['TaskID', 'ProjectID', 'Duration', 'Weight', 'Progress', 'ParentTaskID']:
                tasks_df[col] = 0
            else:
                tasks_df[col] = ""

    for col in ORDER_COLUMNS:
        if col not in orders_df.columns:
            orders_df[col] = ""

    for col in PENDING_WORK_COLUMNS:
        if col not in pending_work_df.columns:
            pending_work_df[col] = "" if col in ['Description', 'Status', 'DueDate'] else 0

    # Convert ID columns to numeric if possible
    for df in (projects_df, tasks_df, orders_df, pending_work_df):
        for c in df.columns:
            if 'ID' in c:
                df[c] = pd.to_numeric(df[c], errors='coerce')


def save_data():
    with pd.ExcelWriter(DATABASE_FILE, engine='openpyxl') as writer:
        projects_df.to_excel(writer, sheet_name='Projects', index=False)
        tasks_df.to_excel(writer, sheet_name='Tasks', index=False)
        orders_df.to_excel(writer, sheet_name='Orders', index=False)
        pending_work_df.to_excel(writer, sheet_name='PendingWork', index=False)


# ------------------------------------------------------------
# AUTO-CALCULATE PROJECT SUB-PROGRESS FROM TASKS
# ------------------------------------------------------------


# ------------------------------------------------------------
# MAIN APPLICATION
# ------------------------------------------------------------
class FullProjectManagerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Full Project Tracking App")
        self.geometry("1400x800")
        self.selected_pending_id = None

        self.selected_project_id = None
        self.figure_canvas = None
        self.orders_tree_context_menu = None

        load_data()
        self.create_tabs()
        self.refresh_project_list()

    # --------------------------------------------------------
    # TABS
    # --------------------------------------------------------


    def update_project_subprogress(self, project_id):
        """
        Update the project's overall progress by averaging sub-progresses.
        """
        global projects_df, tasks_df

        if project_id not in projects_df['ProjectID'].values:
            return

        # Get tasks related to this project
        proj_tasks = tasks_df[tasks_df['ProjectID'] == project_id]

        # If no tasks, set all progress to 0
        if proj_tasks.empty:
            idx = projects_df[projects_df['ProjectID'] == project_id].index
            if not idx.empty:
                i = idx[0]
                for sp in TASK_SUBCATEGORIES.values():
                    projects_df.at[i, sp] = 0
                projects_df.at[i, 'OverallProgress'] = 0
            save_data()
            return

        # Compute sub-progress from related tasks
        subprogress_values = {}
        for cat_name, proj_col in TASK_SUBCATEGORIES.items():
            cat_tasks = proj_tasks[proj_tasks['Category'] == cat_name]
            if cat_tasks.empty:
                subprogress_values[proj_col] = 0.0
            else:
                subprogress_values[proj_col] = cat_tasks['Progress'].mean()

        # Compute overall project progress as an average of sub-progresses
        overall = sum(subprogress_values.values()) / len(TASK_SUBCATEGORIES)

        # Store in projects_df
        idx = projects_df[projects_df['ProjectID'] == project_id].index
        if not idx.empty:
            i = idx[0]
            for col_key, val in subprogress_values.items():
                projects_df.at[i, col_key] = val
            projects_df.at[i, 'OverallProgress'] = overall

        save_data()




    
    def create_tabs(self):
        self.tab_control = ttk.Notebook(self)

        # Projects tab
        self.projects_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.projects_tab, text="Projects")

        # Tasks tab
        self.tasks_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tasks_tab, text="Tasks")

        # Orders tab
        self.orders_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.orders_tab, text="Orders")

        # Reports tab
        self.reports_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.reports_tab, text="Reports")

        self.tab_control.pack(expand=1, fill="both")

        self.build_projects_tab()
        self.build_tasks_tab()
        self.build_orders_tab()
        self.build_reports_tab()

    # --------------------------------------------------------
    # PROJECTS TAB
    # --------------------------------------------------------
    def build_projects_tab(self):
        frame = self.projects_tab

        # Top frame (add project)
        top_frame = tk.LabelFrame(frame, text="Add / Delete Projects", padx=10, pady=10)
        top_frame.pack(fill="x", padx=5, pady=5)

        tk.Label(top_frame, text="Project Name:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.project_name_entry = tk.Entry(top_frame, width=40)
        self.project_name_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Button(top_frame, text="Add Project", command=self.add_project).grid(row=0, column=2, padx=5, pady=5)

        # Project list
        mid_frame = tk.LabelFrame(frame, text="Projects List", padx=10, pady=10)
        mid_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.projects_listbox = tk.Listbox(mid_frame, height=8, exportselection=False)
        self.projects_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.projects_listbox.bind('<<ListboxSelect>>', self.on_project_select)

        scrollbar = tk.Scrollbar(mid_frame, orient="vertical", command=self.projects_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.projects_listbox.config(yscrollcommand=scrollbar.set)

        # Delete Project
        tk.Button(frame, text="Delete Selected Project", command=self.delete_project).pack(pady=5)

    def add_project(self):
        global projects_df
        project_name = self.project_name_entry.get().strip()
        if not project_name:
            messagebox.showwarning("Input Error", "Please enter a project name.")
            return

        if projects_df.empty:
            next_id = 1
        else:
            next_id = projects_df['ProjectID'].max() + 1

        new_row = {
            'ProjectID': next_id,
            'ProjectName': project_name,
            'Notes': "",
            'ElectricalProgress': 0,
            'SSProgress': 0,
            'PlumbingProgress': 0,
            'ScreedProgress': 0,
            'FireSuppressionProgress': 0,
            'OverallProgress': 0,
        }
        projects_df = pd.concat([projects_df, pd.DataFrame([new_row])], ignore_index=True)
        save_data()
        self.refresh_project_list()
        self.project_name_entry.delete(0, tk.END)

    def refresh_project_list(self):
        self.projects_listbox.delete(0, tk.END)
        if projects_df.empty:
            return
        for _, row in projects_df.iterrows():
            pid = int(row['ProjectID'])
            pname = row['ProjectName']
            self.projects_listbox.insert(tk.END, f"{pid}: {pname}")

    def on_project_select(self, event):
        """Update the selected project and refresh all related data."""
        selection = self.projects_listbox.curselection()
        if not selection:
            self.selected_project_id = None
            return

        # ✅ Get the selected project ID
        index = selection[0]
        project_info = self.projects_listbox.get(index)
        project_id = int(project_info.split(":")[0])
        self.selected_project_id = project_id

        # ✅ Refresh tasks & orders for the selected project
        self.refresh_task_list()
        self.refresh_orders_tree()

        # ✅ Refresh pending tasks ONLY if the pending work window is open
        proj_tasks = tasks_df[tasks_df["ProjectID"] == self.selected_project_id]
        if not proj_tasks.empty and hasattr(self, "pending_listbox"):
            first_task_id = proj_tasks.iloc[0]["TaskID"]
            # ✅ Only refresh if pending_listbox still exists
        if hasattr(self, "pending_listbox") and self.pending_listbox.winfo_exists():
            self.refresh_pending_list(first_task_id)

    def delete_project(self):
        global projects_df, tasks_df, orders_df, pending_work_df
        selection = self.projects_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Select a project to delete.")
            return
        confirm = messagebox.askyesno(
            "Confirm Delete",
            "This will remove the project and all associated tasks, orders, and pending work. Proceed?"
        )
        if not confirm:
            return

        index = selection[0]
        project_info = self.projects_listbox.get(index)
        project_id = int(project_info.split(":")[0])

        # Remove associated data
        projects_df = projects_df[projects_df['ProjectID'] != project_id]
        tasks_df = tasks_df[tasks_df['ProjectID'] != project_id]
        orders_df = orders_df[orders_df['ProjectID'] != project_id]
        pending_work_df = pending_work_df[pending_work_df['ProjectID'] != project_id]  # ✅ Remove related pending work

        save_data()

        self.selected_project_id = None
        self.refresh_project_list()
        self.refresh_task_list()
        self.refresh_orders_tree()



    # --------------------------------------------------------
    # TASKS TAB
    # --------------------------------------------------------
    def build_tasks_tab(self):
        frame = self.tasks_tab

        # Label for selected project
        self.selected_project_label_task = tk.Label(frame, text="Selected Project: None")
        self.selected_project_label_task.pack(pady=5)

        add_task_frame = tk.LabelFrame(frame, text="Add Task", padx=10, pady=10)
        add_task_frame.pack(fill="x", padx=5, pady=5)

        tk.Label(add_task_frame, text="Task Name:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.task_name_entry = tk.Entry(add_task_frame, width=40)
        self.task_name_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(add_task_frame, text="Category:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.task_category_var = tk.StringVar()
        self.task_category_combo = ttk.Combobox(
            add_task_frame, textvariable=self.task_category_var,
            values=list(TASK_SUBCATEGORIES.keys()), state="readonly", width=37
        )
        self.task_category_combo.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_task_frame, text="Duration (days):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.task_duration_entry = tk.Entry(add_task_frame, width=10)
        self.task_duration_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_task_frame, text="Progress (%):").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.task_progress_entry = tk.Entry(add_task_frame, width=10)
        self.task_progress_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        

        # Task List
        list_frame = tk.LabelFrame(frame, text="Task List", padx=10, pady=10)
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.task_listbox = tk.Listbox(list_frame, exportselection=False)
        self.task_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        task_scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.task_listbox.yview)
        task_scrollbar.pack(side=tk.RIGHT, fill="y")
        self.task_listbox.config(yscrollcommand=task_scrollbar.set)

        tk.Label(add_task_frame, text="Pending Work:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        tk.Button(add_task_frame, text="Manage Pending Work", command=self.open_pending_work_window).grid(row=4, column=1, padx=5, pady=5, sticky="w")
        tk.Button(add_task_frame, text="Add Task", command=self.add_task).grid(row=5, column=1, padx=5, pady=5, sticky="e")

        # Update progress
        update_frame = tk.LabelFrame(frame, text="Update Task Progress", padx=10, pady=10)
        update_frame.pack(fill="x", padx=5, pady=5)

        tk.Label(update_frame, text="New Progress (%):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.update_progress_entry = tk.Entry(update_frame, width=10)
        self.update_progress_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Button(update_frame, text="Update Progress", command=self.update_task_progress).grid(row=0, column=2, padx=5, pady=5)

        # Delete Task
        tk.Button(frame, text="Delete Selected Task", command=self.delete_task).pack(pady=5)
        tk.Button(frame, text="Edit Pending Items", command=self.edit_pending_items).pack(pady=5)

    def add_task(self):
        global tasks_df
        if self.selected_project_id is None:
            messagebox.showwarning("No Project", "Select a project first.")
            return
        name = self.task_name_entry.get().strip()
        category = self.task_category_var.get().strip()
        dur_str = self.task_duration_entry.get().strip()
        prog_str = self.task_progress_entry.get().strip()

            # Ensure self.task_pending_entry exists
        if hasattr(self, 'task_pending_entry'):
            pending_items = self.task_pending_entry.get().strip()  # Get pending items
        else:
            pending_items = ""  # Default to empty if the entry doesn't exist

        if not name or not category:
            messagebox.showwarning("Input Error", "Task name & category are required.")
            return
        try:
            duration = float(dur_str) if dur_str else 0.0
        except ValueError:
            messagebox.showwarning("Input Error", "Invalid duration.")
            return
        try:
            progress = float(prog_str) if prog_str else 0.0
            if progress < 0 or progress > 100:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Input Error", "Progress must be 0-100.")
            return

        if tasks_df.empty:
            next_tid = 1
        else:
            next_tid = tasks_df['TaskID'].max() + 1

        new_task = {
        'TaskID': next_tid,
        'ProjectID': self.selected_project_id,
        'TaskName': name,
        'Duration': duration,
        'Weight': 0,
        'Progress': progress,
        'ParentTaskID': None,
        'Category': category,
        'PendingItems': pending_items  # Store pending items
    }
        tasks_df = pd.concat([tasks_df, pd.DataFrame([new_task])], ignore_index=True)
        save_data()

        # Update subprogress
        self.update_project_subprogress(self.selected_project_id)

        self.refresh_task_list()

        # Clear fields
        self.task_name_entry.delete(0, tk.END)
        self.task_category_var.set("")
        self.task_duration_entry.delete(0, tk.END)
        self.task_progress_entry.delete(0, tk.END)
        if hasattr(self, 'task_pending_entry'):  # Ensure pending items entry exists
            self.task_pending_entry.delete(0, tk.END)  # Clear pending items field
            
    def open_pending_work_window(self):
        selection = self.task_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Select a task first.")
            return

        # Get selected task ID
        item_str = self.task_listbox.get(selection[0])
        tid = int(item_str.split()[1].replace(":", ""))

        if hasattr(self, 'pending_window') and self.pending_window.winfo_exists():
            self.pending_window.lift()
            return

        self.pending_window = tk.Toplevel(self)
        self.pending_window.title(f"Manage Pending Work - Task {tid}")

        tk.Label(self.pending_window, text="Pending Work Description:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.desc_entry = tk.Entry(self.pending_window, width=50)
        self.desc_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self.pending_window, text="Status:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(self.pending_window, textvariable=self.status_var, 
                                         values=["Pending", "In Progress", "Resolved"], state="readonly")
        self.status_combo.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(self.pending_window, text="Due Date (YYYY-MM-DD):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.due_date_entry = tk.Entry(self.pending_window, width=20)
        self.due_date_entry.grid(row=2, column=1, padx=5, pady=5)

        tk.Button(self.pending_window, text="Add/Update Pending Work", command=lambda: self.add_or_update_pending_work(tid)).grid(row=3, column=1, pady=10, sticky="w")

        # List of pending work
        list_frame = tk.Frame(self.pending_window)
        list_frame.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

        self.pending_listbox = tk.Listbox(list_frame, width=80, height=8)
        self.pending_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.pending_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.pending_listbox.config(yscrollcommand=scrollbar.set)

        self.pending_listbox.bind("<<ListboxSelect>>", lambda event: self.populate_pending_fields(event, tid))


        self.pending_listbox.bind("<Button-1>", lambda event: self.check_deselect(event))  # Detect outside click


        tk.Button(self.pending_window, text="Delete Selected", command=lambda: self.delete_pending_work(tid)).grid(row=5, column=0, columnspan=2, pady=5)

        self.refresh_pending_list(tid)

    def add_or_update_pending_work(self, task_id):
        global pending_work_df

        desc = self.desc_entry.get().strip()
        status = self.status_var.get()
        due_date = self.due_date_entry.get().strip()

        if not desc or not status or not due_date:
            messagebox.showerror("Error", "All fields are required.")
            return  

        if self.selected_project_id is None:
            messagebox.showerror("Error", "No project selected. Please select a project first.")
            return  

        if self.selected_pending_id is not None:
            # Update existing pending task
            pending_work_df.loc[
                (pending_work_df['PendingID'] == self.selected_pending_id) &
                (pending_work_df['ProjectID'] == self.selected_project_id) &
                (pending_work_df['TaskID'] == task_id),
                ['Description', 'Status', 'DueDate']
            ] = [desc, status, due_date]

            self.selected_pending_id = None  # Reset selection after update

        else:
            # Ensure the pending work DataFrame has columns
            if pending_work_df.empty:
                pending_work_df = pd.DataFrame(columns=PENDING_WORK_COLUMNS)

            next_pid = 1 if pending_work_df.empty else pending_work_df['PendingID'].max() + 1

            new_pending = {
                'PendingID': next_pid,
                'TaskID': task_id,
                'ProjectID': self.selected_project_id,  # Assign project ID correctly
                'Description': desc,
                'Status': status,
                'DueDate': due_date
            }

            pending_work_df = pd.concat([pending_work_df, pd.DataFrame([new_pending])], ignore_index=True)

        save_data()
        self.refresh_pending_list(task_id)
        self.clear_pending_fields()


        def check_deselect(self, event):
            """Deselect pending work when clicking outside of a selected item."""
            if not self.pending_listbox.curselection():
                self.selected_pending_id = None
                self.clear_pending_fields()





    def delete_pending_work(self, task_id):
        selection = self.pending_listbox.curselection()
        if not selection:
            messagebox.showerror("Error", "Select a pending work item to delete.")
            return

        confirm = messagebox.askyesno("Confirm Delete", "Delete selected pending work?")
        if not confirm:
            return

        item_str = self.pending_listbox.get(selection[0])
        pid = int(item_str.split()[1].replace(":", ""))

        global pending_work_df
        pending_work_df = pending_work_df[pending_work_df['PendingID'] != pid]

        save_data()
        self.refresh_pending_list(task_id)

        # ✅ Recalculate the task progress after deletion
        #self.update_task_progress_based_on_pending(task_id)


    def update_task_progress_based_on_pending(self, task_id):
        global tasks_df, pending_work_df

        # Get all pending work items for this task
        task_pending_work = pending_work_df[pending_work_df['TaskID'] == task_id]

        # If there are no pending tasks, maintain current progress
        if task_pending_work.empty:
            return  

        total_items = len(task_pending_work)
        resolved_items = len(task_pending_work[task_pending_work['Status'] == "Resolved"])

        # Calculate completion ratio from pending work
        pending_completion_ratio = resolved_items / total_items if total_items > 0 else 1

        # Find the task index in tasks_df
        task_idx = tasks_df[tasks_df['TaskID'] == task_id].index
        if task_idx.empty:
            return  # Task not found

        # Retrieve manual progress input
        manual_progress = tasks_df.at[task_idx[0], 'Progress']

        # Weighted progress calculation (50% manual, 50% pending work)
        updated_progress = (manual_progress * 0.5) + (pending_completion_ratio * 100 * 0.5)

        # Update the task progress
        tasks_df.at[task_idx[0], 'Progress'] = round(updated_progress, 2)

        # ✅ Save the changes
        save_data()

        # ✅ Ensure project progress updates correctly
        project_id = tasks_df.at[task_idx[0], 'ProjectID']

        # ✅ Call the method correctly using `self`
        if hasattr(self, "update_project_subprogress"):
            self.update_project_subprogress(project_id)
        else:
            print("Warning: update_project_subprogress method not found in self.")




    def refresh_pending_list(self, task_id):
        """Refresh the pending work list and ensure the listbox exists."""
        if not hasattr(self, "pending_listbox") or not self.pending_listbox.winfo_exists():
            return  # ✅ Avoid error if pending_listbox does not exist

        self.pending_listbox.delete(0, tk.END)

        if self.selected_project_id is None:
            return  # No project selected, so don't display anything

        # ✅ Ensure it filters correctly by TaskID and ProjectID
        proj_pending = pending_work_df[
            (pending_work_df["TaskID"] == task_id) & 
            (pending_work_df["ProjectID"] == self.selected_project_id)
        ]

        for _, row in proj_pending.iterrows():
            self.pending_listbox.insert(
                tk.END, f"ID {row['PendingID']}: {row['Description']} | {row['Status']} | Due: {row['DueDate']}"
            )






    def add_pending_work(self, task_id, desc_entry, status_var, due_date_entry, top_window):
        """Adds a pending work item to the pending_work_df dataframe."""
        global pending_work_df
        desc = desc_entry.get().strip()
        status = status_var.get()
        due_date = due_date_entry.get().strip()

        if not desc or not status or not due_date:
            messagebox.showerror("Error", "All fields are required.")
            return  # Stop function if input is invalid

        # Generate a new PendingID
        if pending_work_df.empty:
            next_pid = 1
        else:
            next_pid = pending_work_df['PendingID'].max() + 1

        new_pending = {
            'PendingID': next_pid,
            'TaskID': task_id,
            'Description': desc,
            'Status': status,
            'DueDate': due_date
        }

        # Add new pending work to the dataframe
        pending_work_df = pd.concat([pending_work_df, pd.DataFrame([new_pending])], ignore_index=True)
        save_data()  # Save updated data

        messagebox.showinfo("Success", "Pending work added.")
        top_window.destroy()  # Close the popup
        self.open_pending_work_window()  # Refresh the window
    def clear_pending_fields(self):
        """Clears the input fields for pending work."""
        self.desc_entry.delete(0, tk.END)
        self.status_var.set("")
        self.due_date_entry.delete(0, tk.END)


    def populate_pending_fields(self, event, tid):
        """Handles updating the pending fields when an item is selected."""
        selection = self.pending_listbox.curselection()
        if not selection:
            return

        item_str = self.pending_listbox.get(selection[0])
        pending_id = int(item_str.split()[1].replace(":", ""))

        # Find the selected pending work entry
        row = pending_work_df[pending_work_df['PendingID'] == pending_id]
        if not row.empty:
            self.desc_entry.delete(0, tk.END)
            self.desc_entry.insert(0, row.iloc[0]['Description'])

            self.status_var.set(row.iloc[0]['Status'])

            self.due_date_entry.delete(0, tk.END)
            self.due_date_entry.insert(0, row.iloc[0]['DueDate'])

            self.selected_pending_id = pending_id  # Keep track of selected ID




    def edit_pending_items(self):
        selection = self.task_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Select a task to edit pending items.")
            return
        
        item_str = self.task_listbox.get(selection[0])
        tid = int(item_str.split()[1].replace(":", ""))
        
        current_pending = tasks_df.loc[tasks_df['TaskID'] == tid, 'PendingItems'].values[0]
        if pd.isna(current_pending):
            current_pending = ""

        top = tk.Toplevel(self)
        top.title("Edit Pending Items")
        
        tk.Label(top, text="Pending Items:").pack(padx=10, pady=5)
        pending_entry = tk.Text(top, width=50, height=4)
        pending_entry.insert(tk.END, current_pending)
        pending_entry.pack(padx=10, pady=5)
        
        def save_changes():
            new_pending = pending_entry.get("1.0", tk.END).strip()
            tasks_df.loc[tasks_df['TaskID'] == tid, 'PendingItems'] = new_pending
            save_data()
            self.refresh_task_list()
            top.destroy()
        
        tk.Button(top, text="Save", command=save_changes).pack(pady=10)

    def refresh_task_list(self):
        self.task_listbox.delete(0, tk.END)
        if self.selected_project_id is None:
            return
        proj_tasks = tasks_df[tasks_df['ProjectID'] == self.selected_project_id]
        for _, row in proj_tasks.iterrows():
            tid = row['TaskID']
            tname = row['TaskName']
            cat = row['Category']
            dur = row['Duration']
            prog = row['Progress']
            pending = row['PendingItems'] if pd.notna(row['PendingItems']) else ""
            self.task_listbox.insert(
                tk.END,
                f"ID {tid}: {tname} ({cat}) - Dur:{dur} days, {prog}% | Pending: {pending}"
            )

    def update_task_progress(self):
        global tasks_df
        selection = self.task_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Select a task to update.")
            return
        new_prog_str = self.update_progress_entry.get().strip()
        try:
            new_prog = float(new_prog_str)
            if new_prog < 0 or new_prog > 100:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Input Error", "Progress must be 0-100.")
            return

        item_str = self.task_listbox.get(selection[0])
        # e.g. "ID 5: MyTask (Electrical) - Dur:2 days, 50.0%"
        tokens = item_str.split()
        if len(tokens) < 2:
            return
        tid_str = tokens[1]  # "5:"
        tid_str = tid_str.replace(":", "")
        try:
            tid_val = int(tid_str)
        except:
            return

        tasks_df.loc[tasks_df['TaskID'] == tid_val, 'Progress'] = new_prog
        save_data()
        if self.selected_project_id is not None:
            update_project_subprogress(self.selected_project_id)
        self.refresh_task_list()
        self.update_progress_entry.delete(0, tk.END)

    def delete_task(self):
        global tasks_df
        selection = self.task_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Select a task.")
            return
        confirm = messagebox.askyesno("Delete Task", "Are you sure?")
        if not confirm:
            return
        item_str = self.task_listbox.get(selection[0])
        tokens = item_str.split()
        if len(tokens) < 2:
            return
        tid_str = tokens[1].replace(":", "")
        try:
            tid_val = int(tid_str)
        except:
            return
        tasks_df = tasks_df[tasks_df['TaskID'] != tid_val]
        save_data()
        if self.selected_project_id is not None:
            update_project_subprogress(self.selected_project_id)
        self.refresh_task_list()

    # --------------------------------------------------------
    # ORDERS TAB
    # --------------------------------------------------------
    def build_orders_tab(self):
        frame = self.orders_tab
        self.selected_project_label_orders = tk.Label(frame, text="Selected Project: None")
        self.selected_project_label_orders.pack(pady=5)

        add_order_frame = tk.LabelFrame(frame, text="Add New Order", padx=10, pady=5)
        add_order_frame.pack(fill="x", padx=5, pady=5)

        tk.Label(add_order_frame, text="Company:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.company_combobox = ttk.Combobox(add_order_frame, values=COMPANY_NAMES, state="readonly", width=30)
        self.company_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Item Category:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.item_category_combobox = ttk.Combobox(add_order_frame, values=ITEM_CATEGORIES, state="readonly", width=30)
        self.item_category_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Order Status:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.order_status_combobox = ttk.Combobox(add_order_frame, values=ORDER_STATUSES, state="readonly", width=30)
        self.order_status_combobox.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="LPO Status:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.lpo_status_combobox = ttk.Combobox(add_order_frame, values=LPO_STATUSES, state="readonly", width=30)
        self.lpo_status_combobox.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Invoice Status:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.invoice_status_combobox = ttk.Combobox(add_order_frame, values=INVOICE_STATUSES, state="readonly", width=30)
        self.invoice_status_combobox.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Missing Items:").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.missing_items_var = tk.StringVar()
        tk.Entry(add_order_frame, textvariable=self.missing_items_var, width=30)\
            .grid(row=0, column=3, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Delivery Date:").grid(row=1, column=2, padx=5, pady=5, sticky="e")
        self.delivery_date_var = tk.StringVar()
        tk.Entry(add_order_frame, textvariable=self.delivery_date_var, width=30)\
            .grid(row=1, column=3, padx=5, pady=5, sticky="w")

        tk.Label(add_order_frame, text="Installation Date:").grid(row=2, column=2, padx=5, pady=5, sticky="e")
        self.installation_date_var = tk.StringVar()
        tk.Entry(add_order_frame, textvariable=self.installation_date_var, width=30)\
            .grid(row=2, column=3, padx=5, pady=5, sticky="w")

        tk.Button(add_order_frame, text="Add Order", command=self.add_order, bg="green", fg="white")\
            .grid(row=5, column=1, pady=10, sticky="w")
        tk.Button(add_order_frame, text="Delete Order", command=self.delete_order, bg="red", fg="white")\
            .grid(row=5, column=2, pady=10, sticky="w")

        # Add company
        tk.Label(add_order_frame, text="Add New Company:").grid(row=6, column=0, padx=5, pady=5, sticky="e")
        self.new_company_var = tk.StringVar()
        tk.Entry(add_order_frame, textvariable=self.new_company_var, width=30)\
            .grid(row=6, column=1, padx=5, pady=5, sticky="w")
        tk.Button(add_order_frame, text="Add Company", command=self.add_new_company, bg="blue", fg="white")\
            .grid(row=6, column=2, padx=5, pady=5, sticky="w")

        tree_frame = tk.Frame(frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = (
            "OrderID", "Company", "ItemCategory", "OrderStatus",
            "LPOStatus", "Invoice?", "InvoiceStatus",
            "MissingItems", "DeliveryDate", "InstallationDate"
        )
        self.orders_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode="extended")
        for col in columns:
            self.orders_tree.heading(col, text=col)
            self.orders_tree.column(col, width=120, anchor="center")
        self.orders_tree.pack(side=tk.LEFT, fill="both", expand=True)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.orders_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.orders_tree.configure(yscrollcommand=scrollbar.set)

        # Right-click context menu
        self.orders_tree.bind("<Button-3>", self.show_orders_tree_context_menu)

    def update_orders_tab_title(self):
        if self.selected_project_id is not None:
            row = projects_df[projects_df['ProjectID'] == self.selected_project_id]
            if not row.empty:
                pname = row.iloc[0]['ProjectName']
                self.selected_project_label_orders.config(text=f"Selected Project: {pname}")
        else:
            self.selected_project_label_orders.config(text="Selected Project: None")

    def add_new_company(self):
        global COMPANY_NAMES
        new_company = self.new_company_var.get().strip()
        if not new_company:
            messagebox.showerror("Error", "Enter a valid company name.")
            return
        if new_company in COMPANY_NAMES:
            messagebox.showerror("Error", "This company already exists.")
            return
        COMPANY_NAMES.append(new_company)
        self.company_combobox['values'] = COMPANY_NAMES
        self.new_company_var.set('')
        messagebox.showinfo("Success", f"Company '{new_company}' added.")

    def add_order(self):
        global orders_df
        if self.selected_project_id is None:
            messagebox.showerror("Error", "Select a project first.")
            return
        company = self.company_combobox.get()
        item_cat = self.item_category_combobox.get()
        order_stat = self.order_status_combobox.get()
        lpo_stat = self.lpo_status_combobox.get()
        inv_stat = self.invoice_status_combobox.get()

        missing_items = self.missing_items_var.get().strip()
        delivery_date = self.delivery_date_var.get().strip()
        installation_date = self.installation_date_var.get().strip()

        if not item_cat or not order_stat or not lpo_stat or not inv_stat:
            messagebox.showerror("Error", "Fill required order fields.")
            return

        if orders_df.empty:
            next_oid = 1
        else:
            next_oid = orders_df['OrderID'].max() + 1

        new_order = {
            'OrderID': next_oid,
            'ProjectID': self.selected_project_id,
            'Company': company,
            'ItemCategory': item_cat,
            'OrderStatus': order_stat,
            'LPOStatus': lpo_stat,
            'InvoiceCopyPath': "",
            'InvoiceStatus': inv_stat,
            'MissingItems': missing_items,
            'DeliveryDate': delivery_date,
            'InstallationDate': installation_date
        }
        orders_df = pd.concat([orders_df, pd.DataFrame([new_order])], ignore_index=True)
        save_data()
        self.refresh_orders_tree()
        messagebox.showinfo("Success", "Order added.")

        # Clear fields
        self.company_combobox.set("")
        self.item_category_combobox.set("")
        self.order_status_combobox.set("")
        self.lpo_status_combobox.set("")
        self.invoice_status_combobox.set("")
        self.missing_items_var.set("")
        self.delivery_date_var.set("")
        self.installation_date_var.set("")

    def refresh_orders_tree(self):
        self.update_orders_tab_title()
        for row in self.orders_tree.get_children():
            self.orders_tree.delete(row)
        if self.selected_project_id is None:
            return
        project_orders = orders_df[orders_df['ProjectID'] == self.selected_project_id]
        for _, row in project_orders.iterrows():
            inv_uploaded = "Yes" if row['InvoiceCopyPath'] else "No"
            self.orders_tree.insert("", "end", values=(
                row['OrderID'],
                row['Company'],
                row['ItemCategory'],
                row['OrderStatus'],
                row['LPOStatus'],
                inv_uploaded,
                row['InvoiceStatus'],
                row['MissingItems'],
                row['DeliveryDate'],
                row['InstallationDate']
            ))

    def delete_order(self):
        global orders_df
        selection = self.orders_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Select order(s) to delete.")
            return
        confirm = messagebox.askyesno("Confirm", "Delete selected order(s)?")
        if not confirm:
            return
        order_ids = []
        for item in selection:
            oid = self.orders_tree.item(item, "values")[0]
            order_ids.append(int(oid))
        for oid in order_ids:
            orders_df = orders_df[orders_df['OrderID'] != oid]
        save_data()
        self.refresh_orders_tree()
        messagebox.showinfo("Success", f"Deleted {len(selection)} order(s).")

    def show_orders_tree_context_menu(self, event):
        row_id = self.orders_tree.identify_row(event.y)
        if row_id:
            self.orders_tree.selection_set(row_id)
            if not self.orders_tree_context_menu:
                self.orders_tree_context_menu = tk.Menu(self, tearoff=0)
                self.orders_tree_context_menu.add_command(label="Upload Invoice", command=self.upload_invoice)
                self.orders_tree_context_menu.add_command(label="Open Invoice", command=self.open_invoice_copy)
                self.orders_tree_context_menu.add_command(label="Edit Order/LPO Status", command=self.edit_order_lpo_status)
                self.orders_tree_context_menu.add_command(label="Edit Invoice Status", command=self.edit_invoice_status)
                self.orders_tree_context_menu.add_command(label="Edit Additional Fields", command=self.edit_additional_fields)
                self.orders_tree_context_menu.add_command(label="Edit Company", command=self.edit_company)
            self.orders_tree_context_menu.post(event.x_root, event.y_root)

    def get_selected_order_id(self):
        sel = self.orders_tree.selection()
        if not sel:
            return None
        val = self.orders_tree.item(sel[0], "values")[0]
        return int(val)

    def upload_invoice(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        file_path = filedialog.askopenfilename(
            title="Select Invoice Copy",
            filetypes=[("Supported Files", "*.pdf;*.jpg;*.jpeg;*.png"), ("All Files", "*.*")]
        )
        if not file_path:
            return
        global orders_df
        idx = orders_df[orders_df['OrderID'] == oid].index
        if not idx.empty:
            orders_df.at[idx[0], 'InvoiceCopyPath'] = file_path
            save_data()
            self.refresh_orders_tree()
            messagebox.showinfo("Success", "Invoice uploaded.")

    def open_invoice_copy(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        row = orders_df.loc[orders_df['OrderID'] == oid]
        if row.empty:
            return
        path = row.iloc[0]['InvoiceCopyPath']
        if path and os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showerror("Error", "Invoice not found.")

    def edit_order_lpo_status(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        row = orders_df.loc[orders_df['OrderID'] == oid]
        if row.empty:
            return
        old_order_status = row.iloc[0]['OrderStatus']
        old_lpo_status = row.iloc[0]['LPOStatus']

        top = tk.Toplevel(self)
        top.title("Edit Order & LPO Status")

        tk.Label(top, text="Order Status:").grid(row=0, column=0, padx=5, pady=5)
        order_var = tk.StringVar(value=old_order_status)
        order_combo = ttk.Combobox(top, textvariable=order_var, values=ORDER_STATUSES, state="readonly")
        order_combo.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(top, text="LPO Status:").grid(row=1, column=0, padx=5, pady=5)
        lpo_var = tk.StringVar(value=old_lpo_status)
        lpo_combo = ttk.Combobox(top, textvariable=lpo_var, values=LPO_STATUSES, state="readonly")
        lpo_combo.grid(row=1, column=1, padx=5, pady=5)

        def save_changes():
            global orders_df
            orders_df.loc[orders_df['OrderID'] == oid, 'OrderStatus'] = order_var.get()
            orders_df.loc[orders_df['OrderID'] == oid, 'LPOStatus'] = lpo_var.get()
            save_data()
            self.refresh_orders_tree()
            top.destroy()

        tk.Button(top, text="Save Changes", command=save_changes).grid(row=2, column=0, columnspan=2, pady=10)

    def edit_invoice_status(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        row = orders_df.loc[orders_df['OrderID'] == oid]
        if row.empty:
            return
        old_inv_status = row.iloc[0]['InvoiceStatus']

        top = tk.Toplevel(self)
        top.title("Edit Invoice Status")
        tk.Label(top, text="Invoice Status:").grid(row=0, column=0, padx=5, pady=5)
        inv_var = tk.StringVar(value=old_inv_status)
        inv_combo = ttk.Combobox(top, textvariable=inv_var, values=INVOICE_STATUSES, state="readonly")
        inv_combo.grid(row=0, column=1, padx=5, pady=5)

        def save_changes():
            global orders_df
            orders_df.loc[orders_df['OrderID'] == oid, 'InvoiceStatus'] = inv_var.get()
            save_data()
            self.refresh_orders_tree()
            top.destroy()

        tk.Button(top, text="Save Changes", command=save_changes).grid(row=1, column=0, columnspan=2, pady=10)

    def edit_additional_fields(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        row = orders_df.loc[orders_df['OrderID'] == oid]
        if row.empty:
            return
        old_missing = row.iloc[0]['MissingItems']
        old_delivery = row.iloc[0]['DeliveryDate']
        old_installation = row.iloc[0]['InstallationDate']

        top = tk.Toplevel(self)
        top.title("Edit Additional Fields")

        tk.Label(top, text="Missing Items:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        missing_var = tk.StringVar(value=old_missing)
        tk.Entry(top, textvariable=missing_var, width=30).grid(row=0, column=1, padx=5, pady=5)

        tk.Label(top, text="Delivery Date:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        delivery_var = tk.StringVar(value=old_delivery)
        tk.Entry(top, textvariable=delivery_var, width=30).grid(row=1, column=1, padx=5, pady=5)

        tk.Label(top, text="Installation Date:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        install_var = tk.StringVar(value=old_installation)
        tk.Entry(top, textvariable=install_var, width=30).grid(row=2, column=1, padx=5, pady=5)

        def save_changes():
            global orders_df
            orders_df.loc[orders_df['OrderID'] == oid, 'MissingItems'] = missing_var.get()
            orders_df.loc[orders_df['OrderID'] == oid, 'DeliveryDate'] = delivery_var.get()
            orders_df.loc[orders_df['OrderID'] == oid, 'InstallationDate'] = install_var.get()
            save_data()
            self.refresh_orders_tree()
            top.destroy()

        tk.Button(top, text="Save Changes", command=save_changes).grid(row=3, column=0, columnspan=2, pady=10)

    def edit_company(self):
        oid = self.get_selected_order_id()
        if oid is None:
            return
        row = orders_df.loc[orders_df['OrderID'] == oid]
        if row.empty:
            return
        old_company = row.iloc[0]['Company']

        top = tk.Toplevel(self)
        top.title("Edit Company")

        tk.Label(top, text="Company:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        comp_var = tk.StringVar(value=old_company)
        comp_combo = ttk.Combobox(top, textvariable=comp_var, values=COMPANY_NAMES, state="readonly", width=30)
        comp_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        def save_changes():
            global orders_df
            orders_df.loc[orders_df['OrderID'] == oid, 'Company'] = comp_var.get()
            save_data()
            self.refresh_orders_tree()
            top.destroy()

        tk.Button(top, text="Save Changes", command=save_changes).grid(row=1, column=0, columnspan=2, pady=10)

    # --------------------------------------------------------
    # REPORTS TAB
    # --------------------------------------------------------
    def build_reports_tab(self):
        frame = self.reports_tab
        self.selected_project_label_report = tk.Label(frame, text="Selected Project: None")
        self.selected_project_label_report.pack(pady=5)

        tk.Button(frame, text="Generate Project Report", command=self.generate_project_report).pack(pady=5)
        tk.Button(frame, text="Export All Data to Excel", command=self.export_all_data_to_excel).pack(pady=5)

        self.report_charts_frame = tk.Frame(frame)
        self.report_charts_frame.pack(fill="both", expand=True)

    def generate_project_report(self):
        if self.selected_project_id is None:
            messagebox.showwarning("Selection Error", "Select a project first.")
            return
        row = projects_df.loc[projects_df['ProjectID'] == self.selected_project_id]
        if row.empty:
            messagebox.showwarning("No Data", "Project not found.")
            return

        project_name = row.iloc[0]['ProjectName']
        overall_progress = row.iloc[0]['OverallProgress']
        notes = row.iloc[0]['Notes']

        # Gather subprogress
        sub_data = {}
        for cat in TASK_SUBCATEGORIES.values():
            val = row.iloc[0][cat]
            if pd.isna(val):
                val = 0.0
            sub_data[cat] = float(val)

        # tasks for this project
        proj_tasks = tasks_df[tasks_df['ProjectID'] == self.selected_project_id]
        # orders for this project
        proj_orders = orders_df[orders_df['ProjectID'] == self.selected_project_id]

        # We'll create a matplotlib figure with sub-progress (pie) and tasks progress (bar)
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
        fig.suptitle(f"Project: {project_name} (Overall: {overall_progress:.2f}%)")

        # Pie chart of subprogress
        labels = list(sub_data.keys())
        values = list(sub_data.values())
        ax1.pie(values, labels=labels, autopct='%1.1f%%', startangle=140)
        ax1.set_title("Sub-Progress Distribution")

        # bar chart of tasks
        ax2.barh(proj_tasks['TaskName'], proj_tasks['Progress'], color='skyblue')
        ax2.set_xlim(0, 100)
        ax2.set_xlabel("Progress (%)")
        ax2.set_ylabel("Tasks")
        ax2.set_title("Tasks Progress")

        self.display_figure(fig)

        # Also create PDF
        pdf_path = self.create_pdf_report(project_name, overall_progress, notes, sub_data, proj_tasks, proj_orders)
        if pdf_path:
            webbrowser.open_new(pdf_path)

    def display_figure(self, fig):
        if self.figure_canvas:
            self.figure_canvas.get_tk_widget().destroy()
        self.figure_canvas = FigureCanvasTkAgg(fig, master=self.report_charts_frame)
        self.figure_canvas.draw()
        self.figure_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_pdf_report(self, project_name, overall_progress, notes, sub_data, proj_tasks, proj_orders):
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            pdf_path = tmp.name
            tmp.close()

            doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            normal_style = styles['Normal']
            heading_style = styles['Heading2']

            # Title
            elements.append(Paragraph(f"Project Report: {project_name}", title_style))
            elements.append(Spacer(1, 12))

            # Overall progress & notes
            elements.append(Paragraph(f"Overall Progress: {overall_progress:.2f}%", normal_style))
            elements.append(Paragraph(f"Notes: {notes}", normal_style))
            elements.append(Spacer(1, 12))

            # Progress bar
            bar_width = 300
            bar_height = 20
            fill_width = max(0, min(bar_width, bar_width * (overall_progress / 100.0)))
            d = Drawing(bar_width, bar_height)
            d.add(Rect(0, 0, bar_width, bar_height, strokeColor=colors.black, fillColor=colors.lightgrey))
            d.add(Rect(0, 0, fill_width, bar_height, fillColor=colors.green))
            elements.append(Paragraph("Overall Progress Bar:", heading_style))
            elements.append(Spacer(1, 6))
            elements.append(d)
            elements.append(Spacer(1, 12))

            # Sub-progress table
            sub_data_list = [["Sub-Task", "Progress (%)"]]
            for k, v in sub_data.items():
                sub_data_list.append([k, f"{v:.2f}"])
            sub_table = Table(sub_data_list, colWidths=[150, 100])
            sub_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(Paragraph("Sub-Progress Details:", heading_style))
            elements.append(Spacer(1, 6))
            elements.append(sub_table)
            elements.append(Spacer(1, 12))

            # Tasks table
            if not proj_tasks.empty:
                tasks_data_list = [["TaskID", "TaskName", "Category", "Duration", "Progress"]]
                for _, row in proj_tasks.iterrows():
                    tasks_data_list.append([
                        row['TaskID'], row['TaskName'], row['Category'], row['Duration'], f"{row['Progress']}%"
                    ])
                tasks_table = Table(tasks_data_list, repeatRows=1)
                tasks_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ]))
                elements.append(Paragraph("Tasks:", heading_style))
                elements.append(Spacer(1, 6))
                elements.append(tasks_table)
                elements.append(Spacer(1, 12))

                # Add Pending Work for each task
                for _, task_row in proj_tasks.iterrows():
                    task_id = task_row['TaskID']
                    task_name = task_row['TaskName']
                    task_pending = pending_work_df[pending_work_df["TaskID"] == task_id]

                    if not task_pending.empty:
                        elements.append(Paragraph(f"Pending Work for Task: {task_name}", heading_style))
                        pending_data_list = [["PendingID", "Description", "Status", "Due Date"]]
                        for _, p_row in task_pending.iterrows():
                            pending_data_list.append([
                                p_row['PendingID'], p_row['Description'], p_row['Status'], p_row['DueDate']
                            ])
                        pending_table = Table(pending_data_list, repeatRows=1)
                        pending_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ]))
                        elements.append(Spacer(1, 6))
                        elements.append(pending_table)
                        elements.append(Spacer(1, 12))

            # Orders table
            if not proj_orders.empty:
                orders_data_list = [[
                    "OrderID", "Company", "ItemCategory", "OrderStatus",
                    "LPOStatus", "Invoice?", "InvoiceStatus", "MissingItems",
                    "DeliveryDate", "InstallationDate"
                ]]
                for _, rowo in proj_orders.iterrows():
                    inv_up = "Yes" if rowo['InvoiceCopyPath'] else "No"
                    orders_data_list.append([
                        rowo['OrderID'], rowo['Company'], rowo['ItemCategory'], rowo['OrderStatus'],
                        rowo['LPOStatus'], inv_up, rowo['InvoiceStatus'], rowo['MissingItems'],
                        rowo['DeliveryDate'], rowo['InstallationDate']
                    ])
                orders_table = Table(orders_data_list, repeatRows=1)
                orders_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ]))
                elements.append(Paragraph("Orders:", heading_style))
                elements.append(Spacer(1, 6))
                elements.append(orders_table)
            else:
                elements.append(Paragraph("No orders found.", normal_style))

            doc.build(elements)
            return pdf_path
        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to generate PDF: {e}")
            return None


    def export_all_data_to_excel(self):
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not file_path:
                return

            wb = Workbook()

            # Write projects data to "Projects" sheet
            ws_projects = wb.active
            ws_projects.title = "Projects"
            ws_projects.append(PROJECT_COLUMNS)
            for _, row in projects_df.iterrows():
                ws_projects.append([row[col] for col in PROJECT_COLUMNS])

            # Loop through each project and create a separate sheet for tasks & orders
            for _, project_data in projects_df.iterrows():
                project_id = project_data["ProjectID"]
                project_name = project_data["ProjectName"]

                # Ensure the project name is Excel-sheet friendly (no special characters)
                safe_project_name = "".join(c if c.isalnum() or c.isspace() else "_" for c in project_name)
                safe_project_name = safe_project_name[:30]  # Limit to 30 chars

                # Create a sheet for Tasks
                task_sheet = wb.create_sheet(f"{safe_project_name}_Tasks")
                task_sheet.append(TASK_COLUMNS)
                proj_tasks = tasks_df[tasks_df["ProjectID"] == project_id]
                for _, task_row in proj_tasks.iterrows():
                    task_sheet.append([task_row[col] for col in TASK_COLUMNS])

                # Create a sheet for Orders
                order_sheet = wb.create_sheet(f"{safe_project_name}_Orders")
                order_sheet.append(ORDER_COLUMNS)
                proj_orders = orders_df[orders_df["ProjectID"] == project_id]
                for _, order_row in proj_orders.iterrows():
                    order_sheet.append([order_row[col] for col in ORDER_COLUMNS])

                # Create a sheet for Charts
                chart_sheet = wb.create_sheet(f"{safe_project_name}_Charts")

                # Pie Chart for Sub-Progress
                chart_sheet["A1"] = "Sub-Progress Overview"
                categories = list(TASK_SUBCATEGORIES.values())
                values = [project_data.get(cat, 0) for cat in categories]

                # Insert category labels and values
                chart_sheet.append(["Category", "Progress"])
                for cat, val in zip(categories, values):
                    chart_sheet.append([cat, val])

                from openpyxl.chart import PieChart, Reference, BarChart

                pie_chart = PieChart()
                labels = Reference(chart_sheet, min_col=1, min_row=3, max_row=3 + len(categories) - 1)
                data = Reference(chart_sheet, min_col=2, min_row=2, max_row=3 + len(categories) - 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                pie_chart.title = "Sub-Progress Breakdown"
                chart_sheet.add_chart(pie_chart, "D5")

                # Bar Chart for Tasks Progress
                chart_sheet["A20"] = "Task Progress Overview"
                task_names = proj_tasks["TaskName"].tolist()
                task_progress = proj_tasks["Progress"].tolist()

                if task_names:
                    chart_sheet.append(["Task", "Progress"])
                    for name, prog in zip(task_names, task_progress):
                        chart_sheet.append([name, prog])

                    bar_chart = BarChart()
                    labels = Reference(chart_sheet, min_col=1, min_row=22, max_row=22 + len(task_names) - 1)
                    data = Reference(chart_sheet, min_col=2, min_row=21, max_row=22 + len(task_names) - 1)
                    bar_chart.add_data(data, titles_from_data=True)
                    bar_chart.set_categories(labels)
                    bar_chart.title = "Tasks Progress"
                    bar_chart.x_axis.title = "Tasks"
                    bar_chart.y_axis.title = "Progress (%)"
                    chart_sheet.add_chart(bar_chart, "D25")

            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Data exported successfully to {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")



# --------------------------------------------------------
# MAIN
# --------------------------------------------------------
if __name__ == "__main__":
    app = FullProjectManagerApp()
    app.mainloop()
